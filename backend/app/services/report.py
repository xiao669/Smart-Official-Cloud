"""报表服务"""
from datetime import date, timedelta
from io import BytesIO
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook

from app.models.medicine import Medicine
from app.models.inventory import Batch, Transaction, Stocktake
from app.services.warning import WarningService


class ReportService:
    """报表服务类"""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_inventory_report(self) -> list:
        """获取库存报表"""
        result = await self.db.execute(
            select(
                Medicine.id,
                Medicine.name,
                Medicine.code,
                func.sum(Batch.quantity).label("total_quantity"),
                func.count(Batch.id).label("batch_count")
            )
            .join(Batch, Batch.medicine_id == Medicine.id)
            .where(Medicine.is_deleted == False)
            .group_by(Medicine.id)
            .order_by(Medicine.name)
        )
        rows = list(result.all())

        return [
            {
                "medicine_id": r.id,
                "medicine_name": r.name,
                "medicine_code": r.code,
                "total_quantity": r.total_quantity or 0,
                "batch_count": r.batch_count or 0,
            }
            for r in rows
        ]

    async def export_inventory_report(self) -> bytes:
        """导出库存报表"""
        data = await self.get_inventory_report()

        wb = Workbook()
        ws = wb.active
        ws.title = "库存报表"

        headers = ["药品编码", "药品名称", "库存总量", "批次数"]
        ws.append(headers)

        for item in data:
            ws.append([
                item["medicine_code"],
                item["medicine_name"],
                item["total_quantity"],
                item["batch_count"],
            ])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    async def get_expiry_report(self, status: str = "expiring") -> list:
        """获取效期报表"""
        warning_service = WarningService(self.db)
        config = await warning_service.get_config()
        warning_days = config["expiry_warning_days"]

        today = date.today()

        if status == "expired":
            query = select(Batch).where(
                Batch.expiry_date < today,
                Batch.quantity > 0
            )
        else:  # expiring
            query = select(Batch).where(
                Batch.expiry_date >= today,
                Batch.expiry_date <= today + timedelta(days=warning_days),
                Batch.quantity > 0
            )

        query = query.options(selectinload(Batch.medicine)).order_by(Batch.expiry_date)
        result = await self.db.execute(query)
        batches = list(result.scalars().all())

        return [
            {
                "medicine_id": b.medicine_id,
                "medicine_name": b.medicine.name if b.medicine else None,
                "batch_number": b.batch_number,
                "quantity": b.quantity,
                "expiry_date": b.expiry_date.isoformat(),
                "days_until_expiry": (b.expiry_date - today).days,
            }
            for b in batches
        ]

    async def export_expiry_report(self, status: str = "expiring") -> bytes:
        """导出效期报表"""
        data = await self.get_expiry_report(status)

        wb = Workbook()
        ws = wb.active
        ws.title = "效期报表"

        headers = ["药品名称", "批次号", "数量", "有效期", "剩余天数"]
        ws.append(headers)

        for item in data:
            ws.append([
                item["medicine_name"],
                item["batch_number"],
                item["quantity"],
                item["expiry_date"],
                item["days_until_expiry"],
            ])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    async def get_transaction_report(
        self, start_date: date | None = None, end_date: date | None = None
    ) -> list:
        """获取出入库报表"""
        if not start_date:
            start_date = date.today() - timedelta(days=30)
        if not end_date:
            end_date = date.today()

        result = await self.db.execute(
            select(
                func.date(Transaction.created_at).label("date"),
                Transaction.type,
                func.sum(Transaction.quantity).label("total")
            )
            .where(
                func.date(Transaction.created_at) >= start_date,
                func.date(Transaction.created_at) <= end_date
            )
            .group_by(func.date(Transaction.created_at), Transaction.type)
            .order_by(func.date(Transaction.created_at))
        )
        rows = list(result.all())

        # 按日期汇总
        date_map = {}
        for r in rows:
            d = r.date.isoformat() if hasattr(r.date, 'isoformat') else str(r.date)
            if d not in date_map:
                date_map[d] = {"date": d, "inbound_count": 0, "outbound_count": 0}
            if r.type == "inbound":
                date_map[d]["inbound_count"] = r.total
            else:
                date_map[d]["outbound_count"] = r.total

        return list(date_map.values())

    async def export_transaction_report(
        self, start_date: date | None = None, end_date: date | None = None
    ) -> bytes:
        """导出出入库报表"""
        data = await self.get_transaction_report(start_date, end_date)

        wb = Workbook()
        ws = wb.active
        ws.title = "出入库报表"

        headers = ["日期", "入库数量", "出库数量"]
        ws.append(headers)

        for item in data:
            ws.append([
                item["date"],
                item["inbound_count"],
                item["outbound_count"],
            ])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    async def get_stocktake_report(self) -> list:
        """获取盘点报表"""
        result = await self.db.execute(
            select(Stocktake)
            .options(selectinload(Stocktake.creator), selectinload(Stocktake.items))
            .order_by(Stocktake.created_at.desc())
        )
        stocktakes = list(result.scalars().all())

        return [
            {
                "id": s.id,
                "name": s.name,
                "status": s.status,
                "creator_name": s.creator.realname if s.creator else None,
                "created_at": s.created_at.isoformat(),
                "completed_at": s.completed_at.isoformat() if s.completed_at else None,
                "discrepancy_count": sum(1 for item in s.items if item.discrepancy and item.discrepancy != 0),
            }
            for s in stocktakes
        ]
