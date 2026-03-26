"""药品服务"""
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete as sql_delete, func
from openpyxl import Workbook, load_workbook

from app.services.mode_service import ModeService
from app.schemas.medicine import (
    MedicineCreate, MedicineUpdate, MedicineResponse, 
    MedicineListParams, ImportResult, MedicineCategoryResponse
)
from app.schemas.common import PaginatedResponse
from app.core.exceptions import NotFoundError, BusinessError


class MedicineService(ModeService):
    """药品服务类（支持多场景）"""
    
    def __init__(self, db: AsyncSession, mode: str = "medicine"):
        super().__init__(db, mode)
    
    async def list_medicines(self, params: MedicineListParams) -> PaginatedResponse[MedicineResponse]:
        """获取药品列表"""
        # 使用动态模型
        ItemModel = self.ItemModel
        CategoryModel = self.CategoryModel
        
        query = select(ItemModel).where(ItemModel.is_deleted == False)
        
        # 关键词搜索
        if params.keyword:
            keyword_pattern = f"%{params.keyword}%"
            query = query.where(ItemModel.name.ilike(keyword_pattern))
        
        # 分类筛选
        if params.category_id:
            query = query.where(ItemModel.category_id == params.category_id)
        
        # 排序
        if params.sort_by == "name":
            order_col = ItemModel.name
        elif params.sort_by == "created_at":
            order_col = ItemModel.created_at
        else:
            order_col = ItemModel.id
        
        if params.sort_order == "asc":
            query = query.order_by(order_col.asc())
        else:
            query = query.order_by(order_col.desc())
        
        # 统计总数
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await self.db.execute(count_query)
        total = total_result.scalar() or 0
        
        # 分页
        offset = (params.page - 1) * params.page_size
        query = query.offset(offset).limit(params.page_size)
        
        result = await self.db.execute(query)
        items_list = list(result.scalars().all())
        
        # 获取分类信息
        category_ids = [item.category_id for item in items_list]
        if category_ids:
            categories_result = await self.db.execute(
                select(CategoryModel).where(CategoryModel.id.in_(category_ids))
            )
            categories = {c.id: c.name for c in categories_result.scalars().all()}
        else:
            categories = {}
        
        items = []
        for item in items_list:
            response = MedicineResponse(
                id=item.id,
                code=item.code,
                barcode=getattr(item, 'barcode', None),
                name=item.name,
                category_id=item.category_id,
                category_name=categories.get(item.category_id),
                specification=getattr(item, 'specification', None),
                unit=item.unit,
                manufacturer=getattr(item, 'manufacturer', None),
                price=getattr(item, 'price', None),
                description=getattr(item, 'description', None),
                is_deleted=getattr(item, 'is_deleted', False),
                created_at=item.created_at,
                updated_at=getattr(item, 'updated_at', item.created_at)
            )
            items.append(response)
        
        pages = (total + params.page_size - 1) // params.page_size
        
        return PaginatedResponse(
            items=items,
            total=total,
            page=params.page,
            page_size=params.page_size,
            pages=pages
        )
    
    async def get_medicine(self, id: int) -> MedicineResponse:
        """获取药品详情"""
        ItemModel = self.ItemModel
        CategoryModel = self.CategoryModel
        
        result = await self.db.execute(
            select(ItemModel).where(ItemModel.id == id)
        )
        item = result.scalar_one_or_none()
        if not item:
            raise NotFoundError("物品", id)
        
        # 获取分类名称
        category_result = await self.db.execute(
            select(CategoryModel).where(CategoryModel.id == item.category_id)
        )
        category = category_result.scalar_one_or_none()
        
        response = MedicineResponse(
            id=item.id,
            code=item.code,
            barcode=getattr(item, 'barcode', None),
            name=item.name,
            category_id=item.category_id,
            category_name=category.name if category else None,
            specification=getattr(item, 'specification', None),
            unit=item.unit,
            manufacturer=getattr(item, 'manufacturer', None),
            price=getattr(item, 'price', None),
            description=getattr(item, 'description', None),
            is_deleted=getattr(item, 'is_deleted', False),
            created_at=item.created_at,
            updated_at=getattr(item, 'updated_at', item.created_at)
        )
        return response
    
    async def get_medicine_by_barcode(self, barcode: str) -> MedicineResponse:
        """通过条形码获取药品"""
        ItemModel = self.ItemModel
        CategoryModel = self.CategoryModel
        
        result = await self.db.execute(
            select(ItemModel).where(ItemModel.barcode == barcode)
        )
        item = result.scalar_one_or_none()
        if not item:
            raise NotFoundError("物品条形码", barcode)
        
        # 获取分类名称
        category_result = await self.db.execute(
            select(CategoryModel).where(CategoryModel.id == item.category_id)
        )
        category = category_result.scalar_one_or_none()
        
        response = MedicineResponse(
            id=item.id,
            code=item.code,
            barcode=getattr(item, 'barcode', None),
            name=item.name,
            category_id=item.category_id,
            category_name=category.name if category else None,
            specification=getattr(item, 'specification', None),
            unit=item.unit,
            manufacturer=getattr(item, 'manufacturer', None),
            price=getattr(item, 'price', None),
            description=getattr(item, 'description', None),
            is_deleted=getattr(item, 'is_deleted', False),
            created_at=item.created_at,
            updated_at=getattr(item, 'updated_at', item.created_at)
        )
        return response
    
    async def create_medicine(self, data: MedicineCreate) -> MedicineResponse:
        """创建药品"""
        ItemModel = self.ItemModel
        CategoryModel = self.CategoryModel
        
        # 检查编码是否重复
        existing_result = await self.db.execute(
            select(ItemModel).where(ItemModel.code == data.code)
        )
        existing = existing_result.scalar_one_or_none()
        if existing:
            raise BusinessError("DUPLICATE_CODE", f"物品编码 {data.code} 已存在")
        
        # 检查条形码是否重复（如果提供了条形码）
        if hasattr(data, 'barcode') and data.barcode:
            existing_barcode_result = await self.db.execute(
                select(ItemModel).where(ItemModel.barcode == data.barcode)
            )
            existing_barcode = existing_barcode_result.scalar_one_or_none()
            if existing_barcode:
                raise BusinessError("DUPLICATE_BARCODE", f"条形码 {data.barcode} 已被物品「{existing_barcode.name}」使用")
        
        # 检查分类是否存在
        category_result = await self.db.execute(
            select(CategoryModel).where(CategoryModel.id == data.category_id)
        )
        category = category_result.scalar_one_or_none()
        if not category:
            raise NotFoundError("分类", data.category_id)
        
        # 创建物品
        item = ItemModel(
            name=data.name,
            code=data.code,
            barcode=data.barcode if hasattr(data, 'barcode') else None,
            category_id=data.category_id,
            specification=data.specification if hasattr(data, 'specification') else None,
            unit=data.unit,
            manufacturer=data.manufacturer if hasattr(data, 'manufacturer') else None,
            price=data.price if hasattr(data, 'price') else None,
            description=data.description if hasattr(data, 'description') else None
        )
        self.db.add(item)
        await self.db.commit()
        await self.db.refresh(item)
        
        response = MedicineResponse(
            id=item.id,
            code=item.code,
            barcode=getattr(item, 'barcode', None),
            name=item.name,
            category_id=item.category_id,
            category_name=category.name,
            specification=getattr(item, 'specification', None),
            unit=item.unit,
            manufacturer=getattr(item, 'manufacturer', None),
            price=getattr(item, 'price', None),
            description=getattr(item, 'description', None),
            is_deleted=getattr(item, 'is_deleted', False),
            created_at=item.created_at,
            updated_at=getattr(item, 'updated_at', item.created_at)
        )
        return response
    
    async def update_medicine(self, id: int, data: MedicineUpdate) -> MedicineResponse:
        """更新药品"""
        ItemModel = self.ItemModel
        CategoryModel = self.CategoryModel
        
        # 检查物品是否存在
        result = await self.db.execute(
            select(ItemModel).where(ItemModel.id == id)
        )
        item = result.scalar_one_or_none()
        if not item:
            raise NotFoundError("物品", id)
        
        # 检查条形码是否被其他物品使用（如果更新了条形码）
        if hasattr(data, 'barcode') and data.barcode:
            existing_barcode_result = await self.db.execute(
                select(ItemModel).where(ItemModel.barcode == data.barcode, ItemModel.id != id)
            )
            existing_barcode = existing_barcode_result.scalar_one_or_none()
            if existing_barcode:
                raise BusinessError("DUPLICATE_BARCODE", f"条形码 {data.barcode} 已被物品「{existing_barcode.name}」使用")
        
        # 更新字段
        update_data = data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(item, key, value)
        
        await self.db.commit()
        await self.db.refresh(item)
        
        # 获取分类名称
        category_result = await self.db.execute(
            select(CategoryModel).where(CategoryModel.id == item.category_id)
        )
        category = category_result.scalar_one_or_none()
        
        response = MedicineResponse(
            id=item.id,
            code=item.code,
            barcode=getattr(item, 'barcode', None),
            name=item.name,
            category_id=item.category_id,
            category_name=category.name if category else None,
            specification=getattr(item, 'specification', None),
            unit=item.unit,
            manufacturer=getattr(item, 'manufacturer', None),
            price=getattr(item, 'price', None),
            description=getattr(item, 'description', None),
            is_deleted=getattr(item, 'is_deleted', False),
            created_at=item.created_at,
            updated_at=getattr(item, 'updated_at', item.created_at)
        )
        return response
    
    async def delete_medicine(self, id: int) -> bool:
        """删除药品（同时删除关联的库存记录）"""
        from datetime import date, timedelta
        from app.models.inventory import Batch, Transaction
        from app.models.warning import Warning
        from sqlalchemy import delete
        
        ItemModel = self.ItemModel
        
        # 检查物品是否存在
        result = await self.db.execute(
            select(ItemModel).where(ItemModel.id == id)
        )
        item = result.scalar_one_or_none()
        if not item:
            raise NotFoundError("物品", id)
        
        # 药品场景使用原有的 Batch 表（medicine_id 字段）
        if self.mode == 'medicine':
            # 查询该药品的所有库存批次
            batch_result = await self.db.execute(
                select(Batch).where(Batch.medicine_id == id)
            )
            batches = list(batch_result.scalars().all())
            
            # 检查是否有正常状态的库存（非过期、非临期）
            today = date.today()
            expiry_threshold = today + timedelta(days=90)
            
            has_normal_stock = False
            for batch in batches:
                if batch.quantity > 0:
                    if batch.expiry_date > expiry_threshold:
                        has_normal_stock = True
                        break
            
            if has_normal_stock:
                raise BusinessError(
                    "HAS_STOCK",
                    "该物品还有正常库存，请先清空库存后再删除（过期或临期物品可直接删除）"
                )
            
            # 删除关联记录
            await self.db.execute(
                delete(Warning).where(Warning.medicine_id == id)
            )
            
            if batches:
                batch_ids = [batch.id for batch in batches]
                await self.db.execute(
                    delete(Warning).where(Warning.batch_id.in_(batch_ids))
                )
                await self.db.execute(
                    delete(Transaction).where(Transaction.batch_id.in_(batch_ids))
                )
                await self.db.execute(
                    delete(Batch).where(Batch.id.in_(batch_ids))
                )
        
        # 删除物品
        await self.db.execute(
            sql_delete(ItemModel).where(ItemModel.id == id)
        )
        
        await self.db.commit()
        return True
    
    async def list_categories(self) -> list[MedicineCategoryResponse]:
        """获取分类列表"""
        CategoryModel = self.CategoryModel
        
        result = await self.db.execute(
            select(CategoryModel).order_by(CategoryModel.sort_order, CategoryModel.id)
        )
        categories = list(result.scalars().all())
        
        from datetime import datetime
        return [
            MedicineCategoryResponse(
                id=c.id,
                name=c.name,
                description=getattr(c, 'description', None),
                sort_order=getattr(c, 'sort_order', 0),
                created_at=getattr(c, 'created_at', datetime.now())
            )
            for c in categories
        ]
    
    async def export_to_excel(self, params: MedicineListParams) -> bytes:
        """导出药品到 Excel"""
        # 获取所有数据（不分页）
        params.page_size = 10000
        result = await self.list_medicines(params)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "物品列表"
        
        # 表头
        headers = ["物品编码", "物品名称", "分类", "规格", "单位", "生产厂家", "描述"]
        ws.append(headers)
        
        # 数据
        for item in result.items:
            ws.append([
                item.code,
                item.name,
                item.category_name or "",
                item.specification or "",
                item.unit,
                item.manufacturer or "",
                item.description or ""
            ])
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    async def generate_import_template(self) -> bytes:
        """生成导入模板"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "物品导入模板"
        
        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(name="微软雅黑", size=12)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # 表头
        headers = ["物品编码*", "物品名称*", "分类", "规格", "单位*", "生产厂家", "描述"]
        ws.append(headers)
        
        # 设置表头样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        # 添加示例数据
        example_data = [
            ["ITEM001", "示例物品1", "分类1", "规格1", "个", "生产厂家1", "描述1"],
            ["ITEM002", "示例物品2", "分类2", "规格2", "件", "生产厂家2", "描述2"]
        ]
        
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row_data in example_data:
            ws.append(row_data)
            row_num = ws.max_row
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                cell.fill = example_fill
            ws.row_dimensions[row_num].height = 24
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 35
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    async def import_from_excel(self, file_content: bytes) -> ImportResult:
        """从 Excel 导入药品"""
        CategoryModel = self.CategoryModel
        ItemModel = self.ItemModel
        
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active
        
        success_count = 0
        fail_count = 0
        errors = []
        
        # 跳过表头
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # 跳过空行
                continue
            
            try:
                code, name, category_name, specification, unit, manufacturer, description = row[:7]
                
                if not code or not name or not unit:
                    errors.append(f"第 {row_idx} 行: 编码、名称、单位不能为空")
                    fail_count += 1
                    continue
                
                # 检查编码是否重复
                existing_result = await self.db.execute(
                    select(ItemModel).where(ItemModel.code == str(code))
                )
                existing = existing_result.scalar_one_or_none()
                if existing:
                    errors.append(f"第 {row_idx} 行: 物品编码 {code} 已存在")
                    fail_count += 1
                    continue
                
                # 查找或创建分类
                categories_result = await self.db.execute(select(CategoryModel))
                categories = list(categories_result.scalars().all())
                category = next((c for c in categories if c.name == category_name), None)
                if not category and category_name:
                    category = CategoryModel(name=category_name)
                    self.db.add(category)
                    await self.db.flush()
                
                # 创建物品
                item = ItemModel(
                    code=str(code),
                    name=str(name),
                    category_id=category.id if category else 1,
                    specification=str(specification) if specification else None,
                    unit=str(unit),
                    manufacturer=str(manufacturer) if manufacturer else None,
                    description=str(description) if description else None
                )
                self.db.add(item)
                await self.db.flush()
                success_count += 1
                
            except Exception as e:
                errors.append(f"第 {row_idx} 行: {str(e)}")
                fail_count += 1
        
        await self.db.commit()
        
        return ImportResult(
            success_count=success_count,
            fail_count=fail_count,
            errors=errors
        )
